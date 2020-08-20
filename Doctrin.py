from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait

path = "/Users/chandrashekarbasavaraj/Taskoutput/Data.xlsx"
workbook = openpyxl.load_workbook(path)
worksheet = workbook.active

rows = worksheet.max_row
col = worksheet.max_column

driver = webdriver.Chrome(
    executable_path="/Users/chandrashekarbasavaraj/Documents/chromeDriversSelenium/chromedriver-1")
driver.implicitly_wait(30)
driver.maximize_window()
driver.get("https://doctrin.se/")
time.sleep(5)

button = driver.find_element_by_id("catapultCookie")
button.click()

driver.find_element_by_link_text("ENG").click()

careers = driver.find_element_by_id("menu-item-2399")
careers.click()

time.sleep(5)
wait = WebDriverWait(driver, 10)
alert = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='tt-cookie-alert__accept-all']")))
alert.click()

driver.execute_script("window.scrollBy(0,2000)", "")

Coworkers = driver.find_element_by_xpath("//*[@id='section-people']/div/div/a/span")
Coworkers.click()

driver.execute_script("window.scrollBy(0,1000)", "")
time.sleep(3)

eachel = driver.find_element_by_id("jobsite")
names = eachel.find_elements_by_xpath("//*[@class='thumbnail-hover circle']/span[1]")
jobtitle = eachel.find_elements_by_xpath("//*[@class='thumbnail-hover circle']/span[2]")
emp = {}
for i in range(len(jobtitle)):
    if "Project" in jobtitle[i].text:
        emp[names[i].text] = jobtitle[i].text
    if "Engineer" in jobtitle[i].text:
        emp[names[i].text] = jobtitle[i].text
print(emp)

row = 2
col = 1

for key in emp.keys():
    # row += 1
    #   worksheet.write(row, col, key)
    worksheet.cell(row, col, key)
    # for item in emp[key]:
    worksheet.cell(row, col + 1, emp[key])
    row += 1
    col = 1

workbook.save(path)
driver.quit()